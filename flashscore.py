import bs4
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import json
import openpyxl
import time


def start():
    global chrome_driver, delay, lineups, taticas_rosques, jogos_quinados_url
    chrome_driver = webdriver.Chrome("C:\\Users\\pcost\\chromedriver_win32\\chromedriver.exe")
    delay = 10
    with open("lineups.json") as lineups_json:
        lineups = json.load(lineups_json)
    taticas_rosques = {}
    jogos_quinados_url = []


def close():
    global chrome_driver
    chrome_driver.close()


def normalize_players_vector(players, tactic):
    i = 1
    for number_players_row in tactic.split("-"):
        players[i:int(number_players_row)+i] = reversed(players[i:int(number_players_row)+i])
        i += int(number_players_row)


def scrap_league_seasons(country, league, start_year, last_year):
    return [scrap_league_season(country, league, str(year) + "-" + str(year + 1)) for year in range(start_year, last_year + 1)]


def scrap_league_season(country, league, year):
    print("Season: " + year)
    url = "https://www.flashscore.com/football/{}/{}-{}/results/".format(country, league, year)
    chrome_driver.get(url)

    timeout = False
    while not timeout:
        try:
            myElem = WebDriverWait(chrome_driver, delay).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'event__more')))
            print("Page is ready!")
        except TimeoutException:
            print("Loading page took too much time!")
            timeout = True
        else:
            webelement = chrome_driver.find_element_by_class_name("event__more")
            chrome_driver.execute_script("arguments[0].click();", webelement)
            time.sleep(5)

    soup = bs4.BeautifulSoup(chrome_driver.page_source, "html.parser")
    games_table_soup = soup.find("div", class_="sportName soccer")
    games_table_list = games_table_soup.find_all("div", class_="event__match")
    print("Games loaded: " + str(len(games_table_list)))
    matches = []
    match_index = 1
    for game_soup in games_table_list:
        print("Progress: {:0.2f}%".format((match_index / len(games_table_list))*100))
        match_id = game_soup["id"][4:]
        match_url = "https://www.flashscore.com/match/{}/#lineups;1".format(match_id)
        match = scrap_game(match_url)
        if match is None:
            jogos_quinados_url.append(year + ": " + "https://www.flashscore.com/match/{}/#lineups;1".format(match_url))
        else:
            matches.append(match)
        match_index += 1

    return {"Year": year, "Matches": matches}


def scrap_game(url):
    match_data = {}
    chrome_driver.get(url)
    try:
        myElem = WebDriverWait(chrome_driver, delay).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'h-part')))
        myElem = WebDriverWait(chrome_driver, delay).until(
            EC.presence_of_element_located((By.ID, 'parts')))
        myElem = WebDriverWait(chrome_driver, delay).until(
            EC.presence_of_element_located((By.ID, 'h11')))
        myElem = WebDriverWait(chrome_driver, delay).until(
            EC.presence_of_element_located((By.ID, 'a11')))
        # print("Page is ready!")
    except TimeoutException:
        print("Loading page took too much time!")

    soup = bs4.BeautifulSoup(chrome_driver.page_source, "html.parser")
    match_data["Home Team"] = soup.find("div", class_="team-text tname-home").find("a").get_text()
    match_data["Away Team"] = soup.find("div", class_="team-text tname-away").find("a").get_text()

    teams_formation_soup = soup.find_all("td", class_="h-part")
    home_team_formation = teams_formation_soup[0].get_text().replace(" ", "")
    away_team_formation = teams_formation_soup[2].get_text().replace(" ", "")

    home_team_players = [soup.find(id="h" + str(i)).find("a").get_text() for i in range(1, 12)]
    away_team_players = [soup.find(id="a" + str(i)).find("a").get_text() for i in range(1, 12)]

    normalize_players_vector(home_team_players, home_team_formation)
    normalize_players_vector(away_team_players, away_team_formation)

    match_data["Home Vector"] = []

    if home_team_formation not in lineups["tactics"]:
        if home_team_formation in taticas_rosques:
            taticas_rosques[home_team_formation] += 1
        else:
            taticas_rosques[home_team_formation] = 1
        return None
    else:
        i = 0
        for tactic_position in lineups["tactics"][home_team_formation]:
            if tactic_position == "1":
                match_data["Home Vector"].append(home_team_players[i])
                i += 1
            else:
                match_data["Home Vector"].append(tactic_position)

    match_data["Away Vector"] = []

    if away_team_formation not in lineups["tactics"]:
        if away_team_formation in taticas_rosques:
            taticas_rosques[away_team_formation] += 1
        else:
            taticas_rosques[away_team_formation] = 1
        return None
    else:
        i = 0
        for tactic_position in lineups["tactics"][away_team_formation]:
            if tactic_position == "1":
                match_data["Away Vector"].append(away_team_players[i])
                i += 1
            else:
                match_data["Away Vector"].append(tactic_position)

    print("Home team: {} ({}), Away team: {} ({})".format(match_data["Home Team"], home_team_formation, match_data["Away Team"], away_team_formation))

    return match_data


def create_wb(seasons):

    wb = openpyxl.Workbook()
    current_ws = wb.active

    first_season = True

    for season in seasons:
        if first_season:
            first_season = False
            current_ws.title = season["Year"]
        else:
            current_ws = wb.create_sheet(season["Year"])

        current_column = 1
        current_ws.cell(row=1, column=current_column, value="HT")

        current_column += 1
        current_ws.cell(row=1, column=current_column, value="AT")

        for position in lineups["positions"]:
            current_column += 1
            current_ws.cell(row=1, column=current_column, value="H" + position)

        for position in lineups["positions"]:
            current_column += 1
            current_ws.cell(row=1, column=current_column, value="A" + position)

        current_row = 2
        current_column = 1
        for match in season["Matches"]:
            current_ws.cell(row=current_row, column=current_column, value=match["Home Team"])
            current_column += 1
            current_ws.cell(row=current_row, column=current_column, value=match["Away Team"])
            for vector_element in match["Home Vector"]:
                current_column += 1
                current_ws.cell(row=current_row, column=current_column, value=vector_element)
            for vector_element in match["Away Vector"]:
                current_column += 1
                current_ws.cell(row=current_row, column=current_column, value=vector_element)
            current_row += 1
            current_column = 1

    wb.save("lineups.xlsx")
    wb.close()


def write_taticas_rosques():
    open('TaticasRosques.txt', 'w').close()
    with open("TaticasRosques.txt", "a") as taticas_rosques_file:
        for tatica_rosque in taticas_rosques:
            taticas_rosques_file.write(tatica_rosque + ": " + str(taticas_rosques[tatica_rosque]) + "\n")


def write_jogos_quinados():
    open("JogosQuinados.txt", "w").close()
    with open("JogosQuinados.txt", "a") as jogos_quinados_file:
        for jogo_quinado in jogos_quinados_url:
            jogos_quinados_file.write(jogo_quinado + "\n")
